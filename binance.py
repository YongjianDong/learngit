import os
import traceback
import openpyxl
from openpyxl.styles import Alignment
from tkinter import Tk, filedialog, simpledialog
from datetime import datetime

# 在处理交易记录时，目前只对比Pair，导致使用USDT买入， BUSD卖出的同一货币算不出成本

# 启动程序后让用户选择快照日期，并存储到snapshot_date变量中
Tk().withdraw()
snapshot_date = simpledialog.askstring("选择利润计算截止日期", "请输入日期（格式：YYYYMMDD）")

if snapshot_date:
    # 创建新的Excel文件
    file_name = f"Binance_Profit{snapshot_date}.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active

    # 设置列头
    sheet["A1"] = "UTC"
    sheet["B1"] = "TOKEN"
    sheet["C1"] = "Profit"
    sheet["D1"] = "Cost"
    sheet["E1"] = "Deposit Required"
    #sheet["F1"] = "Deposit record UTC"

    # 设置列宽
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 15

    # 允许用户选择transaction 与 Deposit类型的Excel文件
    prompt = "请选择要处理的Binance成交订单记录文件"
    transaction_file_path = filedialog.askopenfilename(title=prompt, filetypes=[("Excel Files", "*.xlsx")])
    #transaction_file_path ='/Users/ydong/Library/Mobile Documents/com~apple~CloudDocs/010工作台/加密货币入金记录/Binance_transctions_test2021.xlsx'
    

    print(transaction_file_path)
        
    transaction_directory = os.path.dirname(transaction_file_path)
    new_file_path = os.path.join(transaction_directory, file_name)

    prompt = "请选择要处理的Binance Deposit记录文件"
    deposit_file_path = filedialog.askopenfilename(title=prompt, filetypes=[("Excel Files", "*.xlsx")])
    #deposit_file_path = '/Users/ydong/Library/Mobile Documents/com~apple~CloudDocs/010工作台/加密货币入金记录/test_deposite_history_Binance_2021.xlsx'
    print(deposit_file_path)

    

    if  transaction_file_path and deposit_file_path :
        
        try:
            # 加载Transaction文件
            transaction_wb = openpyxl.load_workbook(transaction_file_path)
            transaction_sheet = transaction_wb['Processed-2']

            # Load Deposite File
            deposit_wb = openpyxl.load_workbook(deposit_file_path)
            deposit_sheet = deposit_wb.active
            

            # 创建2个字典来存储depoist中的token数量 以及 BUY记录中的token数量。
            deposit_amounts = {}

            buy_amounts = {}

            # Load buy_row[4] to Dictionary token_amount from transaction file ,  key is UTC, value is Executed.
            for buy_row in transaction_sheet.iter_rows(min_row=2, values_only=True):
                    if buy_row[2] == 'BUY':
                        # found buying records of sold tokens
                        #print(buy_row)
                        buy_amounts[buy_row[0]]=buy_row[4]

            # Load deposit amount to Dictionary deposit_amount from Deposit file
            for deposit_row in deposit_sheet.iter_rows(min_row=2, values_only=True):
                deposit_amounts[deposit_row[0]] = deposit_row[3]
        

            # 遍历transaction文件中的每一行记录
            for row_number, transaction_row in enumerate(transaction_sheet.iter_rows(min_row=2, values_only=True), start=2):

                #print(transaction_row)
                pair_value = transaction_row[1]
                side_value = transaction_row[2]
                if pair_value is None:
                    print('reach to end of sheet')
                    break

                if side_value == 'BUY':
                    continue
                
                pair_lenth = len(pair_value)
                
                token_check = pair_value[-4:]
                
                if token_check == 'USDT' or token_check == 'USDC' or token_check == 'BUSD' :
                    to_token = pair_value[:(pair_lenth-4)]
                    from_token = pair_value[-4:]
                elif token_check.endswith('ETH'):
                    to_token = pair_value[:(pair_lenth-3)]
                    from_token = pair_value[-3:]
                elif token_check.endswith('BTC'):
                    to_token = pair_value[:(pair_lenth-3)]
                    from_token = pair_value[-3:]
                else :
                    print('unexpected pair',pair_value)
                    
                #发现SELL记录，在结果表中增加一行，记录UTC，TOKEN，卖掉的token数量，收入（AMOUNT-FEE），买这些token的成本，收益。
                new_UTC = transaction_row[0]
                #new_token below is new_pair, not token only string, need to cut to token string only when search the deposit file.
                new_token = transaction_row[9]
                new_executed = transaction_row[4]
                new_income = transaction_row[5]
                new_deposit_required = 'Yes'
                new_cost = 0.0
                new_profit = 0.0

                
                # seeking BUY records to find cost and update tokens's sellable amount， 每次都从文件最初开始遍历，确保先买后卖都关系.
                # 遍历到卖记录之前，如果没有发现足够数量到token， 那么就是来自外部， 需要查Deposit记录， 而且也应该在卖记录之前，之后到无效。
                
                for buy_row in transaction_sheet.iter_rows(min_row=2, values_only=True):
                    if buy_row[2] == 'BUY' and buy_row[9] == to_token and buy_row[0] < transaction_row[0]:
                        # found buying records of sold tokens
                        if buy_amounts[buy_row[0]] <= 0 :  #This buy record had been consumed.
                            continue
                        else :
                            if new_executed > buy_amounts[buy_row[0]]:
                                new_cost += buy_row[3]*buy_amounts[buy_row[0]]
                                #if to_token == 'BTC':
                                #   print(' part new cost for BTC sell record', transaction_row[0], 'is', new_cost)
                                new_executed -=buy_amounts[buy_row[0]]
                                buy_amounts[buy_row[0]] = 0
                            else :  # there is enough tokens in buy record
                                #if to_token == 'BTC':
                                #   print('Token', new_token,'found enough in transaction record, buy row ', buy_row[0],buy_row[1],buy_row[3],buy_row[4],buy_row[8],buy_amounts[buy_row[0]],buy_row)
                                #   print('new_cost  and new profit before add is', new_cost, new_profit)
                                buy_amounts[buy_row[0]] -=  new_executed
                                new_cost += (new_executed * buy_row[3] + (new_executed/buy_row[4])*buy_row[8])
                                new_profit = transaction_row[5]-transaction_row[8] - new_cost
                                #if to_token == 'BTC':
                                #    print('transaction row 5, 8 is', transaction_row[5],transaction_row[8])
                                #    print('new_cost and new profit after calculate', new_cost, new_profit)
                                new_deposit_required = 'No'
                                new_row =(new_UTC,new_token, new_profit, new_cost, new_deposit_required)
                                sheet.append(new_row)
                                new_cost = 0.0
                                break
                    else :
                        if buy_row[0] >= new_UTC:  
                            #This token is from out side, need to check deposit record
                            for deposit_row in deposit_sheet.iter_rows(min_row=2, values_only=True):
                                #print('deposit_row is',deposit_row[0])
                                if to_token == deposit_row[1]:
                                    # Found deposit record.
                                    #print('Found deposit record, deposit_row 0 is', to_token,deposit_row[0],'transaction row 0 is', transaction_row[0])
                                    if deposit_amounts[deposit_row[0]] <= 0: # This deposit record had been consummed
                                        continue
                                    else :
                                        if new_executed > deposit_amounts[deposit_row[0]]: # Found part of tokens
                                            print('Token new executed is', to_token, new_executed, 'deposit_amounts is', deposit_row[0], deposit_amounts[deposit_row[0]])
                                            new_cost += deposit_row[4]*deposit_amounts[deposit_row[0]]
                                            new_executed -= deposit_amounts[deposit_row[0]]
                                            deposit_amounts[deposit_row[0]] = 0
                                        else :
                                            #There is engouh token to sell
                                            print('the deposit record has enough tokens is', to_token, deposit_row[0],deposit_row[3])
                                            if deposit_row[5] > 0 and deposit_row[4] > 0 :
                                                new_cost += new_executed/(deposit_amounts[deposit_row[0]])*deposit_row[5]
                                                deposit_amounts[deposit_row[0]] -= new_executed
                                                new_profit = transaction_row[5]-transaction_row[8] - new_cost
                                                new_deposit_required = 'Done'
                                            else :
                                                #需要成本数据， Unit cost and total cost
                                                print('need cost info')
                                            new_row = (new_UTC,to_token, new_profit, new_cost, new_deposit_required)
                                            sheet.append(new_row)
                                            break
                            break
                        else :
                            continue
                        
                if new_deposit_required == 'Yes':
                        # No token deposit found or cost info missing.
                        print(to_token, new_executed,'is not found  or cost info missing')
                        new_profit = 0
                        new_row = (new_UTC,to_token, new_profit, new_cost, new_deposit_required)
                        sheet.append(new_row)


               
            print(" Transactions 处理完成！")
            
        except Exception as e:
             # 获取异常的追踪信息
            traceback.print_tb(e.__traceback__)
            print(f"Transactions 处理过程中出现错误：{str(e)}")


        # 设置单元格对齐方式
            
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # 保存Binance_Snapshot文件
        wb.save(new_file_path)

    else:
        print("未选择Deposit或Withdraw或transactions文件。")
else:
    print("未选交易截止日期。")
